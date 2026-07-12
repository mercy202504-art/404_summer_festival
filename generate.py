from __future__ import annotations

import base64
import os
import time
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook


# ==========================================
# 404 MEMORY FACTORY 設定
# ==========================================

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "memory_master.xlsx"
OUTPUT_DIR = BASE_DIR / "images" / "memory"

SHEET_NAME = "Memory Master"

# 今回生成する最大枚数
MAX_GENERATIONS = 999

# 画像生成設定
IMAGE_MODEL = "gpt-image-2"
IMAGE_SIZE = "1024x1536"
IMAGE_QUALITY = "low"
IMAGE_FORMAT = "webp"
IMAGE_COMPRESSION = 85

# API連続実行の間隔
WAIT_SECONDS = 2


def text(value: object) -> str:
    """Excelの値を安全に文字列へ変換する。"""
    if value is None:
        return ""
    return str(value).strip()


def normalize_no(value: object) -> str:
    """1、01、001などを001形式へ統一する。"""
    raw = text(value)

    try:
        return str(int(float(raw))).zfill(3)
    except (ValueError, TypeError):
        return raw.zfill(3)


def main() -> None:
    # --------------------------------------
    # APIキー読込
    # --------------------------------------
    load_dotenv(BASE_DIR / ".env", override=True)

    api_key = os.getenv("OPENAI_API_KEY")

    if not api_key:
        raise RuntimeError(
            ".envからOPENAI_API_KEYを読み込めませんでした。"
        )

    # --------------------------------------
    # ファイル確認
    # --------------------------------------
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excelファイルが見つかりません：{EXCEL_FILE}"
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # --------------------------------------
    # Excel読込
    # --------------------------------------
    workbook = load_workbook(EXCEL_FILE)

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"シート「{SHEET_NAME}」が見つかりません。"
        )

    sheet = workbook[SHEET_NAME]

    # 1行目から列名を取得
    headers = {
        text(cell.value): cell.column
        for cell in sheet[1]
        if text(cell.value)
    }

    required_columns = [
        "No",
        "Title",
        "ImagePrompt",
        "Filename",
        "Generate",
        "GenerationStatus",
        "Notes",
    ]

    missing_columns = [
        name
        for name in required_columns
        if name not in headers
    ]

    if missing_columns:
        raise ValueError(
            "Excelに必要な列がありません："
            + ", ".join(missing_columns)
        )

    client = OpenAI(api_key=api_key)

    generated_count = 0
    skipped_count = 0
    failed_count = 0

    print()
    print("========================================")
    print("404 MEMORY FACTORY")
    print(f"今回の最大生成枚数：{MAX_GENERATIONS}")
    print("========================================")
    print()

    # --------------------------------------
    # 各行を順番に処理
    # --------------------------------------
    for row_number in range(2, sheet.max_row + 1):
        memory_no = normalize_no(
            sheet.cell(row_number, headers["No"]).value
        )

        title = text(
            sheet.cell(row_number, headers["Title"]).value
        )

        prompt = text(
            sheet.cell(
                row_number,
                headers["ImagePrompt"]
            ).value
        )

        filename = text(
            sheet.cell(
                row_number,
                headers["Filename"]
            ).value
        )

        generate_flag = text(
            sheet.cell(
                row_number,
                headers["Generate"]
            ).value
        ).upper()

        status_cell = sheet.cell(
            row_number,
            headers["GenerationStatus"]
        )

        notes_cell = sheet.cell(
            row_number,
            headers["Notes"]
        )

        current_status = text(status_cell.value).upper()

        # Generate列がYES以外なら対象外
        if generate_flag != "YES":
            continue

        # 生成済みのExcel行は飛ばす
        if current_status == "GENERATED":
            continue

        if not filename:
            filename = f"{memory_no}.webp"

        output_file = OUTPUT_DIR / filename

        # 既に画像が存在する場合
        if output_file.exists():
            print(
                f"[SKIP] No.{memory_no} "
                f"{title}：画像ファイルが存在します。"
            )

            status_cell.value = "GENERATED"
            notes_cell.value = "既存画像を確認してGENERATEDへ更新"
            skipped_count += 1

            workbook.save(EXCEL_FILE)
            continue

        # 今回の上限に達したら終了
        if generated_count >= MAX_GENERATIONS:
            break

        if not prompt:
            print(
                f"[FAILED] No.{memory_no}："
                "ImagePromptが空です。"
            )

            status_cell.value = "FAILED"
            notes_cell.value = "ImagePromptが空です"
            failed_count += 1

            workbook.save(EXCEL_FILE)
            continue

        print("----------------------------------------")
        print(f"生成中：No.{memory_no}")
        print(f"タイトル：{title}")
        print(f"保存先：{output_file}")
        print("----------------------------------------")

        # 処理開始状態を記録
        status_cell.value = "PROCESSING"
        notes_cell.value = "画像生成処理中"
        workbook.save(EXCEL_FILE)

        rarity = text(
            sheet.cell(
                row_number,
                headers["Rarity"]
            ).value
        ).upper()

        quality = "high" if rarity == "RARE" else "low"

        try:
            result = client.images.generate(
                model=IMAGE_MODEL,
                prompt=prompt,
                size=IMAGE_SIZE,
                quality=quality,
                output_format=IMAGE_FORMAT,
                output_compression=IMAGE_COMPRESSION,
                n=1,
            )

            image_base64 = result.data[0].b64_json

            if not image_base64:
                raise RuntimeError(
                    "画像データが返されませんでした。"
                )

            image_bytes = base64.b64decode(image_base64)
            output_file.write_bytes(image_bytes)

            status_cell.value = "GENERATED"
            notes_cell.value = (
                f"生成完了：{filename}"
            )

            generated_count += 1

            print(
                f"[SUCCESS] No.{memory_no}を生成しました。"
            )

        except KeyboardInterrupt:
            status_cell.value = "PENDING"
            notes_cell.value = "ユーザー操作で中断"
            workbook.save(EXCEL_FILE)

            print()
            print("生成を中断しました。")
            print("次回は続きから再開できます。")
            return

        except Exception as error:
            status_cell.value = "FAILED"
            notes_cell.value = (
                f"{type(error).__name__}: {str(error)[:300]}"
            )

            failed_count += 1

            print(
                f"[FAILED] No.{memory_no}"
            )
            print(error)

        # 1件ごとにExcelへ保存
        workbook.save(EXCEL_FILE)

        if generated_count < MAX_GENERATIONS:
            time.sleep(WAIT_SECONDS)

    # --------------------------------------
    # 最終保存
    # --------------------------------------
    workbook.save(EXCEL_FILE)

    print()
    print("========================================")
    print("今回の処理が終了しました。")
    print(f"新規生成：{generated_count}件")
    print(f"既存スキップ：{skipped_count}件")
    print(f"失敗：{failed_count}件")
    print(f"保存先：{OUTPUT_DIR}")
    print("========================================")


if __name__ == "__main__":
    main()