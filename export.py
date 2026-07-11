from pathlib import Path
import json
from openpyxl import load_workbook


# export.py と同じフォルダを基準にする
BASE_DIR = Path(__file__).resolve().parent

EXCEL_FILE = BASE_DIR / "memory_master.xlsx"
OUTPUT_FILE = BASE_DIR / "memory-data.js"


def text(value) -> str:
    """Excelの空欄を安全に文字列へ変換する。"""
    if value is None:
        return ""
    return str(value).strip()


def memory_number(value) -> str:
    """1 → 001 のように3桁へ整える。"""
    raw = text(value)

    try:
        return f"{int(float(raw)):03d}"
    except (ValueError, TypeError):
        return raw.zfill(3)


def main() -> None:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excelが見つかりません：{EXCEL_FILE.name}"
        )

    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active

    # 1行目を列名として取得
    headers = {
        text(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if text(cell.value)
    }

    required = [
        "No",
        "MemoryID",
        "Word1",
        "Word2",
        "Title",
        "Poem",
        "Mood",
        "Composition",
        "Lighting",
        "ImagePrompt",
        "Filename",
        "Status",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise ValueError(
            "Excelに必要な列がありません："
            + ", ".join(missing)
        )

    cards = {}

    for row in range(2, sheet.max_row + 1):
        word1 = text(sheet.cell(row, headers["Word1"]).value)
        word2 = text(sheet.cell(row, headers["Word2"]).value)

        # 空の行は無視
        if not word1 or not word2:
            continue

        no = memory_number(
            sheet.cell(row, headers["No"]).value
        )

        filename = text(
            sheet.cell(row, headers["Filename"]).value
        )

        key = f"{word1}_{word2}"

        cards[key] = {
            "no": no,
            "memoryId": text(
                sheet.cell(row, headers["MemoryID"]).value
            ),
            "word1": word1,
            "word2": word2,
            "title": text(
                sheet.cell(row, headers["Title"]).value
            ),
            "poem": text(
                sheet.cell(row, headers["Poem"]).value
            ),
            "mood": text(
                sheet.cell(row, headers["Mood"]).value
            ),
            "composition": text(
                sheet.cell(row, headers["Composition"]).value
            ),
            "lighting": text(
                sheet.cell(row, headers["Lighting"]).value
            ),
            "prompt": text(
                sheet.cell(row, headers["ImagePrompt"]).value
            ),
            "image": f"images/memory/{filename}",
            "status": text(
                sheet.cell(row, headers["Status"]).value
            ),
        }

    json_data = json.dumps(
        cards,
        ensure_ascii=False,
        indent=2
    )

    javascript = (
        "// memory_master.xlsx から自動生成したファイルです。\n"
        "// このファイルは直接編集せず、Excelを編集してください。\n\n"
        f"const MEMORY_CARDS = {json_data};\n"
    )

    OUTPUT_FILE.write_text(
        javascript,
        encoding="utf-8"
    )

    print("--------------------------------")
    print("404観測記録を書き出しました。")
    print(f"件数：{len(cards)}")
    print(f"出力：{OUTPUT_FILE.name}")
    print("--------------------------------")


if __name__ == "__main__":
    main()